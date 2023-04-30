from googleapiclient.discovery import build

import openpyxl
import os
import subprocess
import xlsxwriter


## Fetches project IDs to which credentials have permissions to access
def generate_gcp_projects_list(file_path):

    print("Running scripts to generate the list GCP Projects")

    row = 0
    column = 0

    ## Creating a workbook for the mentioned Path and adding a sheet inside it
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet("GCP_projects")

    ## Getting the list of projects using gcloud command and storing it in temp.txt file
    subprocess.check_output("gcloud projects list > ./artifacts/temp.txt", shell=True)
    
    ## Adding a heading to the first column
    worksheet.write(row,column,"Project_ID")
    row +=1

    ## Reading list of projects from temp file and writing it into excel
    with open('./artifacts/temp.txt','r') as file:
        flag = 0
        for line in file:      
            word = line.split()  
            if str(word[0]) != "PROJECT_ID":
                worksheet.write(row,column,str(word[0]))
                column += 1

    ## Deleting the temp file
    os.remove("./artifacts/temp.txt")
    ## Closing the workbook to save the file
    workbook.close()
    print(f"Excel with list of GCP Projects is created at {file_path}\n")


## Reads data from excel and stores it in a LIST
def get_gcp_projects(file_path):

    project_list = []
    ## Creating a dataFrame and storing the GCP_projects excel data
    dataframe = openpyxl.load_workbook(file_path).active

    ## Traversing through dataFrame and storing each project ID in LIST
    for row in range(1,dataframe.max_row):
        for col in dataframe.iter_cols(1,dataframe.max_column):
            if col[row].value is not None:
                project_list.append(col[row].value)

    ## Return the list of Project IDs
    return project_list
